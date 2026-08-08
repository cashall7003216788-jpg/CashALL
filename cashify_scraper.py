"""
Cashify Laptop Price Scraper v5
- Collects /sell-old-laptop/used-* links per brand page
- Global dedup: each model URL assigned to the FIRST brand page that shows it
- Brand-affinity filter: model slug must NOT clearly belong to a different brand
- Parallel price fetching with 2 worker pages
"""

import re
import asyncio
from playwright.async_api import async_playwright, TimeoutError as PWTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LOG_FILE    = r"c:\Users\DELL\OneDrive\Desktop\CashALL\scraper_log.txt"
OUTPUT_FILE = r"c:\Users\DELL\OneDrive\Desktop\CashALL\Cashify_Laptop_Models_Prices.xlsx"

BRAND_URLS = [
    ("Xiaomi",       "https://www.cashify.in/sell-old-laptop/sell-xiaomi"),
    ("Apple",        "https://www.cashify.in/sell-old-laptop/sell-apple"),
    ("Samsung",      "https://www.cashify.in/sell-old-laptop/sell-samsung"),
    ("Realme",       "https://www.cashify.in/sell-old-laptop/sell-realme"),
    ("Lenovo",       "https://www.cashify.in/sell-old-laptop/sell-lenovo"),
    ("Nokia",        "https://www.cashify.in/sell-old-laptop/sell-nokia"),
    ("Dell",         "https://www.cashify.in/sell-old-laptop/sell-dell"),
    ("HP-Compaq",    "https://www.cashify.in/sell-old-laptop/sell-hp-compaq"),
    ("Asus",         "https://www.cashify.in/sell-old-laptop/sell-asus"),
    ("LG",           "https://www.cashify.in/sell-old-laptop/sell-lg"),
    ("Acer",         "https://www.cashify.in/sell-old-laptop/sell-acer"),
    ("Microsoft",    "https://www.cashify.in/sell-old-laptop/sell-microsoft"),
    ("MSI",          "https://www.cashify.in/sell-old-laptop/sell-msi"),
    ("AVITA",        "https://www.cashify.in/sell-old-laptop/sell-avita"),
    ("Other Laptop", "https://www.cashify.in/sell-old-laptop/sell-other-laptop"),
]

# Slug keywords that strongly indicate a specific brand
# Used to filter out cross-brand featured links
BRAND_SLUGS = {
    "Xiaomi":       ["mi-", "xiaomi", "redmibook"],
    "Apple":        ["macbook", "apple"],
    "Samsung":      ["galaxy", "samsung"],
    "Realme":       ["realme"],
    "Lenovo":       ["lenovo", "thinkpad", "ideapad", "yoga", "legion", "thinkbook"],
    "Nokia":        ["nokia"],
    "Dell":         ["dell", "vostro", "inspiron", "xps", "alienware", "latitude", "precision"],
    "HP-Compaq":    ["hp", "pavilion", "probook", "elitebook", "spectre", "envy", "omen", "stream", "compaq"],
    "Asus":         ["asus", "zenbook", "vivobook", "rog", "tuf", "expertbook"],
    "LG":           ["lg-", "gram"],
    "Acer":         ["acer", "aspire", "swift", "nitro", "predator", "spin", "extensa"],
    "Microsoft":    ["microsoft", "surface"],
    "MSI":          ["msi"],
    "AVITA":        ["avita"],
    "Other Laptop": [],  # accepts anything not matched above
}

# Build reverse map: slug_keyword -> brand
KEYWORD_TO_BRAND = {}
for brand, keywords in BRAND_SLUGS.items():
    for kw in keywords:
        KEYWORD_TO_BRAND[kw] = brand


def slug_brand(url_slug: str) -> str:
    """Return the brand this URL slug most likely belongs to, or '' if unknown."""
    slug_lower = url_slug.lower()
    for kw, brand in KEYWORD_TO_BRAND.items():
        if kw in slug_lower:
            return brand
    return ""


_log_fh = open(LOG_FILE, "w", encoding="utf-8", buffering=1)

def log(msg: str):
    print(msg)
    _log_fh.write(msg + "\n")
    _log_fh.flush()


def parse_price(text: str):
    """Extract 'Get Upto Rs. X' price. Returns int (500-500000) or None."""
    m = re.search(r"upto\s*(?:rs\.?\s*)?([\d,]+)", text, re.IGNORECASE)
    if m:
        val = int(m.group(1).replace(",", ""))
        if 500 <= val <= 500000:
            return val
    m2 = re.search(r"\u20b9\s*([\d,]+)", text)
    if m2:
        val = int(m2.group(1).replace(",", ""))
        if 500 <= val <= 500000:
            return val
    m3 = re.search(r"(?:rs\.?|inr)\s*([\d,]+)", text, re.IGNORECASE)
    if m3:
        val = int(m3.group(1).replace(",", ""))
        if 500 <= val <= 500000:
            return val
    return None


async def scroll_page(page, steps=8, pause=0.5):
    for _ in range(steps):
        prev = await page.evaluate("document.body.scrollHeight")
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await asyncio.sleep(pause)
        curr = await page.evaluate("document.body.scrollHeight")
        if curr == prev:
            break


async def get_model_links(page, brand_url: str, brand: str, global_seen: set):
    """
    Collect model links from brand page.
    - Only URLs with /sell-old-laptop/used- pattern
    - Skip URLs already claimed by a previous brand (global_seen)
    - Skip URLs whose slug clearly belongs to a different brand
    """
    for attempt in range(3):
        try:
            await page.goto(brand_url, wait_until="networkidle", timeout=50000)
            break
        except Exception:
            try:
                await page.goto(brand_url, wait_until="domcontentloaded", timeout=50000)
                break
            except Exception as e:
                if attempt == 2:
                    log(f"  FAILED to load: {e}")
                    return []
                await asyncio.sleep(3)

    await asyncio.sleep(2)
    await scroll_page(page)

    all_links = await page.evaluate("""
        () => Array.from(document.querySelectorAll('a[href]')).map(a => ({
            href: a.href,
            text: (a.innerText || a.textContent || '').trim()
        }))
    """)

    results = []
    for item in all_links:
        href = (item.get("href") or "").strip().split("?")[0].split("#")[0].rstrip("/")
        text = (item.get("text") or "").replace("\n", " ").strip()

        if "/sell-old-laptop/used-" not in href:
            continue
        if "cashify.in" not in href:
            continue

        # Skip nested paths (config/question pages)
        after_used = href.split("/sell-old-laptop/used-")[-1]
        if "/" in after_used:
            continue

        # Already claimed by another brand
        if href in global_seen:
            continue

        # Brand-affinity check: slug should match THIS brand or be unrecognized
        detected = slug_brand(after_used)
        if detected and detected != brand:
            # This slug clearly belongs to a different brand — skip
            log(f"  SKIP (belongs to {detected}): {href}")
            continue

        global_seen.add(href)

        # Derive clean name from link text or URL slug
        if not text or len(text) < 2:
            text = after_used.replace("-", " ").title()

        results.append((text, href))

    return results


async def get_price(page, url: str, retries=2):
    """Open model page and extract Get-Upto price."""
    for attempt in range(retries + 1):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=45000)
            await asyncio.sleep(2)
            await scroll_page(page, steps=5, pause=0.4)

            body = await page.inner_text("body")
            price = parse_price(body)
            if price:
                return price

            for sel in ["[class*='price']", "[class*='Price']", "[class*='upto']",
                        "[class*='amount']", "h1", "h2", "h3", "strong", "b"]:
                try:
                    els = await page.query_selector_all(sel)
                    for el in els:
                        t = await el.inner_text()
                        p = parse_price(t)
                        if p:
                            return p
                except Exception:
                    pass

        except PWTimeout:
            if attempt < retries:
                await asyncio.sleep(3)
                continue
        except Exception as e:
            if attempt < retries:
                await asyncio.sleep(3)
                continue
            log(f"    ERROR {url}: {e}")
    return "N/A"


async def scrape_all():
    results = {}
    global_seen = set()  # URLs already assigned to a brand

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled"]
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
            locale="en-IN",
        )
        await context.route(
            "**/*.{png,jpg,jpeg,gif,webp,svg,woff,woff2,ttf,otf,mp4,mp3}",
            lambda r: r.abort()
        )

        pages = [await context.new_page() for _ in range(3)]

        for brand, brand_url in BRAND_URLS:
            log(f"\n=== {brand} ===")
            models = await get_model_links(pages[0], brand_url, brand, global_seen)
            log(f"  Found {len(models)} model(s)")

            if not models:
                results[brand] = []
                continue

            brand_data = []
            for i in range(0, len(models), 2):
                batch = models[i:i+2]
                tasks = [get_price(pages[j+1], url) for j, (_, url) in enumerate(batch)]
                prices = await asyncio.gather(*tasks)
                for (name, url), price in zip(batch, prices):
                    log(f"    [{price}] {name}")
                    brand_data.append((name, price, url))

            results[brand] = brand_data

        await browser.close()

    return results


# ── Excel writer ───────────────────────────────────────────────────────────────

def write_excel(results, path):
    wb = openpyxl.Workbook()

    HDR_FILL  = PatternFill("solid", fgColor="1F3864")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL  = PatternFill("solid", fgColor="D9E1F2")
    NRM_FILL  = PatternFill("solid", fgColor="FFFFFF")
    SUM_FILL  = PatternFill("solid", fgColor="E2EFDA")
    SUM_FONT  = Font(bold=True, color="375623", size=11)
    LINK_FONT = Font(color="1155CC", underline="single", size=10)
    BD = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    def hrow(ws, headers, row):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BD
        ws.row_dimensions[row].height = 22

    def drow(ws, row, vals, alt=False):
        fill = ALT_FILL if alt else NRM_FILL
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.fill = fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = BD

    def cw(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    all_rows = [(b, m, p, u) for b, rows in results.items() for m, p, u in rows]
    total = len(all_rows)
    found = sum(1 for *_, p, _ in all_rows if p != "N/A")
    na    = total - found

    ws_m = wb.active
    ws_m.title = "MASTER"

    for i, (lbl, val) in enumerate([("Total Models", total),
                                     ("Prices Found", found),
                                     ("Prices N/A",   na)], 1):
        for col, v in enumerate([lbl, val], 1):
            c = ws_m.cell(row=i, column=col, value=v)
            c.fill = SUM_FILL; c.font = SUM_FONT
            c.alignment = Alignment(vertical="center"); c.border = BD

    hrow(ws_m, ["Brand", "Model / Series", "Cashify Price (Rs)", "Cashify Model URL"], 5)
    ri = 6
    for brand, rows in results.items():
        for name, price, url in rows:
            pv = price if isinstance(price, int) else "N/A"
            drow(ws_m, ri, [brand, name, pv, url], alt=(ri % 2 == 0))
            if isinstance(pv, int):
                ws_m.cell(row=ri, column=3).number_format = '#,##0'
            c = ws_m.cell(row=ri, column=4)
            c.hyperlink = url; c.font = LINK_FONT
            ri += 1

    cw(ws_m, [15, 40, 20, 60])
    ws_m.freeze_panes = "A6"

    for brand, rows in results.items():
        ws = wb.create_sheet(title=brand[:31])
        hrow(ws, ["Model / Series", "Cashify Price (Rs)", "Cashify Model URL"], 1)
        for i, (name, price, url) in enumerate(rows, 2):
            pv = price if isinstance(price, int) else "N/A"
            drow(ws, i, [name, pv, url], alt=(i % 2 == 0))
            if isinstance(pv, int):
                ws.cell(row=i, column=2).number_format = '#,##0'
            c = ws.cell(row=i, column=3)
            c.hyperlink = url; c.font = LINK_FONT
        cw(ws, [40, 20, 60])
        ws.freeze_panes = "A2"

    wb.save(path)
    log(f"\nSaved: {path}")
    log(f"Total: {total} | Found: {found} | N/A: {na}")


async def main():
    log("=== Cashify Laptop Scraper v5 ===")
    results = await scrape_all()
    write_excel(results, OUTPUT_FILE)
    _log_fh.close()

if __name__ == "__main__":
    asyncio.run(main())
